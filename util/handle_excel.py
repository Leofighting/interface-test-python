# -*- coding:utf-8 -*-
__author__ = "leo"

import os

import openpyxl

base_path = os.path.dirname(os.getcwd())


# open_excel = openpyxl.load_workbook(base_path + "/case/imooc_cases.xlsx")
# sheet_name = open_excel.sheetnames
# excel_value = open_excel[sheet_name[0]]
# print(excel_value)
# print(excel_value.cell(1, 1).value)
# print(excel_value.max_row)


class HandleExcel:
    def load_excel(self):
        """加载 Excel 表格"""
        open_excel = openpyxl.load_workbook(base_path + "/case/imooc_cases.xlsx")
        return open_excel

    def get_sheet_data(self, index=None):
        """加载所有 sheet 的内容"""
        if not index:
            index = 0
        sheet_name = self.load_excel().sheetnames
        data = self.load_excel()[sheet_name[index]]
        return data

    def get_cell_value(self, row, col, index=None):
        """获取某一单元格内容"""
        data = self.get_sheet_data(index).cell(row=row, column=col).value
        return data

    def get_rows(self, index=None):
        return self.get_sheet_data(index).max_row

    def get_row_value(self, row, index=None):
        """获取某一行内容"""
        row_list = []
        for i in self.get_sheet_data(index)[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self, row, col, value):
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row, col, value)
        wb.save(base_path + "/case/imooc_cases.xlsx")


excel_data = HandleExcel()

if __name__ == '__main__':
    handle = HandleExcel()
    handle.excel_write_data(2, 12, "通过")
